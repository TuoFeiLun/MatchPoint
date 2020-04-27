from TransToBaidu import TransToBaiduC
from openpyxl import load_workbook
from openpyxl import Workbook
import SolveEquation
import math
import ConstantParameters
import datetime
'''

model to wgs ， wgs to Baidu mercator 

baidu latitude longitude to baidu mercator

calculate close distance 

'''
# only one Regulator xy data ，  to  baidu09,to baidu mercator


def trans_regulator_to_baidu(standar_regulator: list, standard_wkid):
    transbdc = TransToBaiduC(ConstantParameters.global_ak_code, standard_wkid)

    return_wgs = transbdc.ModelXY_2_WGS(standar_regulator[1], standar_regulator[2])

    return_baidu09mc = transbdc.wgs84_to_bd09mc(return_wgs[0], return_wgs[1])

    standar_regulator.append(return_baidu09mc[0])
    standar_regulator.append(return_baidu09mc[1])
    return standar_regulator

def read_model_regulator(filepath: str):
    wb = load_workbook(filename=filepath)
    ws_1 = wb.active

    all_reg_list = []
    maxrow = ws_1.max_row
    maxrow += 1
    # read reg x y   to list
    for i in range(2, maxrow):
        name = "A" + str(i)
        xcor = 'B' + str(i)
        ycor = 'C' + str(i)
        descrp = 'D' + str(i)
        # 节点名称 x y 描述
        one_list = [ws_1[name].value, ws_1[xcor].value, ws_1[ycor].value, ws_1[descrp].value]

        all_reg_list.append(one_list)

    return all_reg_list

def read_transed_model_data(filepath):
    wb = load_workbook(filename=filepath)
    ws_1 = wb.active

    all_reg_list = []
    maxrow = ws_1.max_row+1
    # read reg x y   to list
    for i in range(1, maxrow):
        name = "A" + str(i)
        xcor = 'B' + str(i)
        ycor = 'C' + str(i)
        descrp = 'D' + str(i)
        transed_x = 'E' +str(i)
        transed_y = 'F' + str(i)
        # 节点名称 x y 描述
        one_list = [ws_1[name].value, ws_1[xcor].value, ws_1[ycor].value, ws_1[descrp].value, ws_1[transed_x].value, ws_1[transed_y].value]

        all_reg_list.append(one_list)

    return all_reg_list

def trans_regulator_to_baidu_special_coordinate(one_reg_list: list, modelwkid: int, result1: list):
    # 注意此时不是标准系直接转 ，对于非国际标准系转换那么需要求解转换参数后，将其转换国际标准再转。
    transToStandardWkidXY = SolveEquation.CalTransformedPoint(result1, one_reg_list[1], one_reg_list[2])
    uns_transbaduc = TransToBaiduC(ConstantParameters.global_ak_code,ConstantParameters.trans_wkid)
    return_wgs = uns_transbaduc.ModelXY_2_WGS(transToStandardWkidXY[0], transToStandardWkidXY[1])

    return_baidu09mc = uns_transbaduc.wgs84_to_bd09mc(return_wgs[0], return_wgs[1])

    one_reg_list.append(return_baidu09mc[0])  # 直接将tuple写入excel 时，会报错
    one_reg_list.append(return_baidu09mc[1])
    return one_reg_list


def save_transed(alldata: list , save_path: str):
    wb_reg = Workbook(write_only=True)
    ws_reg_save = wb_reg.create_sheet()
    for i in alldata:

        ws_reg_save.append(i)
    # 在这里修改路径

    indx = save_path.rfind('.xlsx')

    now_time = datetime.datetime.now()

    str_now_time = now_time.strftime("%Y_%m_%d_%H_%M_%S")
    new_path = save_path[0:indx] + str_now_time + "时间转换后的"+ ".xlsx"

    wb_reg.save(new_path)


def trans_user_to_baidu(user_filepath : str):
    '''
    :param user_filepath:
    :param usersave_path:
    :return:
    '''
    wb_user = load_workbook(filename=user_filepath)
    wsuser = wb_user.active
    user_list = []
    maxrow = wsuser.max_row+1
    trans_user = TransToBaiduC(ConstantParameters.global_ak_code, ConstantParameters.trans_wkid)
    for i in range(2,maxrow):
        name = "A" + str(i)
        xycor = "B" + str(i)
        one_list = [wsuser[name].value, wsuser[xycor].value]
        user_list.append(one_list)

    user_list_len = len(user_list)
    print(user_list)
    print(user_list_len)
    for i in range(0, user_list_len):
        xycor = user_list[i][1]
        if not xycor:   # 跳过空值
            print(xycor)
            continue

        XY_split = xycor.split(',')
        xcor = float(XY_split[0])
        ycor = float(XY_split[1])
        print(xcor,ycor)
        xyBaiduMercator = trans_user.bd09_to_bd09mc(xcor, ycor)
        user_list[i].append(xyBaiduMercator[0])
        user_list[i].append(xyBaiduMercator[1])
        # 存储转化后的数据

    save_transed(user_list, user_filepath)
    return user_list

def calculate_distance(regulator : tuple ,user : tuple) -> float:

    delta_x = abs(regulator[0] - user[0])
    delta_y = abs(regulator[1] - user[1])

    distance = math.sqrt(delta_x**2 + delta_y**2)
    return distance


def read_user_baidu_mc_data(filepath):
    wb = load_workbook(filename=filepath)
    ws_1 = wb.active
    all_user_baidu_mc_data = []
    maxrow = ws_1.max_row + 1
    # read reg x y   to list
    for i in range(1, maxrow):
        name = "A" + str(i)
        user_lnglat = 'B' + str(i)
        xcor = 'C' + str(i)
        ycor = 'D' + str(i)

        one_list = [ws_1[name].value, ws_1[user_lnglat].value, ws_1[xcor].value, ws_1[ycor].value]
        all_user_baidu_mc_data.append(one_list)

    return all_user_baidu_mc_data
